# 程式輔助分析.py
# 1) 先把「示波器電壓量測.xlsx」「示波器頻率量測.xlsx」補齊手動/自動誤差與平均百分誤差 -> *_result.xlsx
# 2) 再從兩個 *_result.xlsx 抓出「電壓(3個波形) + 頻率(3個波形)」共 6 個平均百分誤差
# 3) 画「同一張」長條圖：每個波形一組(Manual/Auto 兩根)，共 6 組

from openpyxl import load_workbook
import os
import math
import numpy as np
import matplotlib.pyplot as plt

plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "Microsoft YaHei", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False

def percent_error(theory, measured):
    if theory == 0:
        return None
    return abs((theory - measured) / theory) * 100

def is_num(x):
    return isinstance(x, (int, float)) and not (isinstance(x, float) and (math.isnan(x) or math.isinf(x)))

def process_file(input_file):
    output_file = input_file.replace(".xlsx", "_result.xlsx")
    wb = load_workbook(input_file)

    for ws in wb.worksheets:
        manual_errors = []
        auto_errors = []

        for row in range(1, ws.max_row + 1):
            theory = ws.cell(row, 1).value   # A欄 理論值/文字
            manual = ws.cell(row, 2).value   # B欄 手動
            auto   = ws.cell(row, 4).value   # D欄 自動

            # 平均百分誤差列
            if isinstance(theory, str) and "平均百分誤差" in theory:
                if manual_errors:
                    ws.cell(row, 3).value = round(sum(manual_errors)/len(manual_errors), 2)
                if auto_errors:
                    ws.cell(row, 5).value = round(sum(auto_errors)/len(auto_errors), 2)
                manual_errors = []
                auto_errors = []
                continue

            # 數據列：A 要是數字
            if is_num(theory):
                if is_num(manual):
                    err_m = percent_error(float(theory), float(manual))
                    ws.cell(row, 3).value = round(err_m, 2)
                    manual_errors.append(err_m)

                if is_num(auto):
                    err_a = percent_error(float(theory), float(auto))
                    ws.cell(row, 5).value = round(err_a, 2)
                    auto_errors.append(err_a)

    wb.save(output_file)
    print(f"完成處理：{output_file}")
    return output_file

def extract_3_block_avgs(result_xlsx, prefix):
    """
    從 *_result.xlsx 擷取三個波形區塊的平均百分誤差（C=手動平均、E=自動平均）
    假設同一張 sheet 內依序出現 3 次「平均百分誤差」列（正弦/方波/三角波）
    回傳 list: [(label, manual_avg, auto_avg), ...] 共 3 筆
    """
    wb = load_workbook(result_xlsx, data_only=True)
    ws = wb.active  # 你的檔通常只有一張表；若有多張可改成迴圈取你要的那張
    out = []

    wave_names = ["正弦波", "方波", "三角波"]
    idx = 0

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "平均百分誤差" in a:
            manual_avg = ws.cell(r, 3).value  # C
            auto_avg   = ws.cell(r, 5).value  # E

            name = wave_names[idx] if idx < 3 else f"Block{idx+1}"
            label = f"{prefix}-{name}"  # 例：電壓-正弦波
            out.append((label, float(manual_avg), float(auto_avg)))
            idx += 1
            if idx >= 3:
                break

    if len(out) != 3:
        raise RuntimeError(f"{result_xlsx} 擷取到 {len(out)} 個平均百分誤差列，請確認表格是否有 3 個區塊。")

    return out

def plot_6_groups(all6, out_png="Manual_vs_Auto_6Avgs.png"):
    """
    all6: list of (label, manual_avg, auto_avg) length=6
    一張圖：6 組，每組 2 根柱(Manual/Auto)
    """
    labels = [x[0] for x in all6]
    manual = np.array([x[1] for x in all6], dtype=float)
    auto   = np.array([x[2] for x in all6], dtype=float)

    x = np.arange(len(labels))
    w = 0.35

    plt.figure(figsize=(12, 4.5))
    plt.bar(x - w/2, manual, width=w, label="Manual")
    plt.bar(x + w/2, auto,   width=w, label="Auto")

    plt.xticks(x, labels, rotation=20, ha="right")
    plt.ylabel("平均百分誤差 (%)")
    plt.title("示波器：電壓與頻率｜手動 vs 自動（6 個平均百分誤差比較）")
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()
    print(f"已輸出圖表：{out_png}")

# ------------------ main ------------------
files = ["示波器電壓量測.xlsx", "示波器頻率量測.xlsx"]
result_files = []

for f in files:
    if os.path.exists(f):
        result_files.append(process_file(f))
    else:
        raise FileNotFoundError(f"找不到檔案：{f}")

# 擷取 6 個平均百分誤差（電壓3 + 頻率3）
voltage_3 = extract_3_block_avgs("示波器電壓量測_result.xlsx", "電壓")
freq_3    = extract_3_block_avgs("示波器頻率量測_result.xlsx", "頻率")
all6 = voltage_3 + freq_3

# 畫在同一張圖
plot_6_groups(all6, out_png="Manual_vs_Auto_6Avgs.png")