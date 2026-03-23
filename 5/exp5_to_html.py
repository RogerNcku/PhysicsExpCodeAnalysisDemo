# -*- coding: utf-8 -*-
from pathlib import Path
import json

import pandas as pd
import numpy as np
from openpyxl import load_workbook


def pick_data_sheet(wb):
    """
    自動選最可能的資料工作表
    """
    best = None
    best_score = -1

    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        header_text = " ".join([str(x) for x in headers if x is not None])

        score = 0
        for kw in ["次數", "X(格數)", "T(s)", "U0(V)", "Q(coulomb)", "區間", "個數"]:
            if kw in header_text:
                score += 1

        if score > best_score:
            best_score = score
            best = ws.title

    return best


def sheet_to_dataframe(ws):
    """
    將 worksheet 轉成 DataFrame
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("工作表是空的")

    header = [str(x).strip() if x is not None else f"col{i+1}" for i, x in enumerate(rows[0])]
    data = rows[1:]
    return pd.DataFrame(data, columns=header)


def clean_numeric(series):
    """
    將欄位轉成數值，無法轉換者設為 NaN
    """
    return pd.to_numeric(series, errors="coerce")


def build_main_table(df):
    """
    建立主資料表，並補算 S、V1、Q
    """
    rename_map = {
        "次數": "trial",
        "X(格數)": "X",
        "S(m)": "S",
        "T(s)": "T",
        "V1(m/s) ": "V1",
        "V1(m/s)": "V1",
        "U0(V)": "U0",
        "Q(coulomb) ": "Q",
        "Q(coulomb)": "Q",
    }

    cols = [c for c in df.columns if c in rename_map]
    if not cols:
        raise ValueError("找不到主資料表欄位，請確認 Excel 表頭是否正確。")

    main = df[cols].copy()
    main = main.rename(columns=rename_map)

    for col in ["trial", "X", "S", "T", "V1", "U0", "Q"]:
        if col in main.columns:
            main[col] = clean_numeric(main[col])

    # 補算 S
    if "S" not in main.columns and "X" in main.columns:
        main["S"] = main["X"] * 0.000382
    elif "S" in main.columns and "X" in main.columns:
        main["S"] = main["S"].fillna(main["X"] * 0.000382)

    # 補算 V1
    if "V1" not in main.columns and {"S", "T"}.issubset(main.columns):
        main["V1"] = main["S"] / main["T"]
    elif "V1" in main.columns and {"S", "T"}.issubset(main.columns):
        main["V1"] = main["V1"].fillna(main["S"] / main["T"])

    # 補算 Q
    if "Q" not in main.columns and {"V1", "U0"}.issubset(main.columns):
        main["Q"] = 1.945e-10 * np.power(main["V1"], 1.5) / main["U0"]
    elif "Q" in main.columns and {"V1", "U0"}.issubset(main.columns):
        main["Q"] = main["Q"].fillna(1.945e-10 * np.power(main["V1"], 1.5) / main["U0"])

    # 清理缺值
    need_cols = [c for c in ["X", "T", "U0", "Q"] if c in main.columns]
    main = main.dropna(subset=need_cols).reset_index(drop=True)

    # 補 trial
    if "trial" not in main.columns:
        main.insert(0, "trial", np.arange(1, len(main) + 1))
    else:
        main["trial"] = main["trial"].fillna(pd.Series(np.arange(1, len(main) + 1), index=main.index))

    return main


def build_distribution(main):
    """
    建立 Q 區間統計表，並加上總數
    """
    step = 0.1e-19
    upper = 5.0e-19
    bins = np.arange(step, upper + step, step)

    counts = []
    labels = []
    q = main["Q"].to_numpy()

    prev = 0.0
    for b in bins:
        count = int(np.sum((q > prev) & (q <= b)))
        labels.append(f"{b:.1E}")
        counts.append(count)
        prev = b

    gt = int(np.sum(q > upper))
    labels.append(">5.0E-19")
    counts.append(gt)

    dist = pd.DataFrame({
        "區間": labels,
        "個數": counts
    })

    total_count = int(dist["個數"].sum())
    total_row = pd.DataFrame([{
        "區間": "總數",
        "個數": total_count
    }])

    dist = pd.concat([dist, total_row], ignore_index=True)
    return dist


def estimate_e(main):
    """
    估計基本電荷 e
    """
    dist = build_distribution(main)

    # 排除最後的總數列與 >5.0E-19
    core = dist[~dist["區間"].isin([">5.0E-19", "總數"])].copy()
    core["edge"] = core["區間"].str.replace("E", "e", regex=False).astype(float)

    candidates = core[(core["edge"] >= 0.8e-19) & (core["個數"] == core["個數"].max())]
    if len(candidates) > 0:
        return float(candidates.iloc[0]["edge"])

    return float(main["Q"].median())


def make_plotly_html(main, dist):
    """
    產生 Plotly 互動圖
    """
    q = main["Q"].to_list()
    v1 = main["V1"].to_list()
    u0 = main["U0"].to_list()
    t = main["T"].to_list()
    trial = main["trial"].astype(int).to_list()

    # 圖表不含總數列
    dist_plot = dist[dist["區間"] != "總數"].copy()
    bar_x = dist_plot["區間"].to_list()
    bar_y = dist_plot["個數"].astype(int).to_list()

    return f"""
<script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
<div class="grid">
  <div id="chart_hist" class="chart"></div>
  <div id="chart_qtrial" class="chart"></div>
  <div id="chart_v1u0" class="chart"></div>
  <div id="chart_tq" class="chart"></div>
</div>
<script>
const trial = {json.dumps(trial)};
const q = {json.dumps(q)};
const v1 = {json.dumps(v1)};
const u0 = {json.dumps(u0)};
const t = {json.dumps(t)};
const barX = {json.dumps(bar_x)};
const barY = {json.dumps(bar_y)};

Plotly.newPlot('chart_hist', [{{
    type:'bar',
    x:barX,
    y:barY,
    hovertemplate:'區間: %{{x}}<br>個數: %{{y}}<extra></extra>'
}}], {{
    title:'N-Q 分布圖',
    xaxis:{{title:'Q 區間'}},
    yaxis:{{title:'個數 N'}}
}}, {{responsive:true}});

Plotly.newPlot('chart_qtrial', [{{
    type:'scatter',
    mode:'lines+markers',
    x:trial,
    y:q,
    hovertemplate:'次數: %{{x}}<br>Q: %{{y:.3e}} C<extra></extra>'
}}], {{
    title:'各次測得的電量 Q',
    xaxis:{{title:'次數'}},
    yaxis:{{title:'Q (C)'}}
}}, {{responsive:true}});

Plotly.newPlot('chart_v1u0', [{{
    type:'scatter',
    mode:'markers',
    x:v1,
    y:u0,
    hovertemplate:'V1: %{{x:.3e}} m/s<br>U0: %{{y:.1f}} V<extra></extra>'
}}], {{
    title:'V1 與 U0 關係',
    xaxis:{{title:'V1 (m/s)'}},
    yaxis:{{title:'U0 (V)'}}
}}, {{responsive:true}});

Plotly.newPlot('chart_tq', [{{
    type:'scatter',
    mode:'markers',
    x:t,
    y:q,
    hovertemplate:'T: %{{x:.2f}} s<br>Q: %{{y:.3e}} C<extra></extra>'
}}], {{
    title:'T 與 Q 關係',
    xaxis:{{title:'T (s)'}},
    yaxis:{{title:'Q (C)'}}
}}, {{responsive:true}});
</script>
"""


def df_to_html_table(df, float_formats=None):
    """
    DataFrame 轉 HTML，並讓總數列加粗
    """
    show = df.copy()

    if float_formats:
        for col, fmt in float_formats.items():
            if col in show.columns:
                show[col] = show[col].map(lambda x: fmt.format(x) if pd.notna(x) else "")

    html = show.to_html(index=False, classes="table", border=0, escape=False)

    html = html.replace("<td>總數</td>", "<td><b>總數</b></td>")

    # 讓最後一格 150 也加粗
    import re
    html = re.sub(
        r"(<td><b>總數</b></td>\s*<td>)(\d+)(</td>)",
        r"\1<b>\2</b>\3",
        html
    )

    return html


def main():
    xlsx_path = Path("exp5.xlsx")
    out_html = Path("exp5_report.html")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = pick_data_sheet(wb)
    if sheet_name is None:
        raise ValueError("找不到適合的工作表。")

    ws = wb[sheet_name]
    df = sheet_to_dataframe(ws)

    main_df = build_main_table(df)
    dist_df = build_distribution(main_df)
    e_est = estimate_e(main_df)

    summary = {
        "資料筆數": len(main_df),
        "Q平均值(C)": float(main_df["Q"].mean()),
        "Q中位數(C)": float(main_df["Q"].median()),
        "Q最小值(C)": float(main_df["Q"].min()),
        "Q最大值(C)": float(main_df["Q"].max()),
        "估計基本電荷e(C)": float(e_est),
    }

    summary_html = "".join(
        f"<tr><th>{k}</th><td>{v:.3e}</td></tr>" if isinstance(v, float)
        else f"<tr><th>{k}</th><td>{v}</td></tr>"
        for k, v in summary.items()
    )
    purpose_html = """
    <p>
    本實驗所產生的 HTML 報告主要分為數個區塊，用以完整呈現密立根油滴實驗的數據分析結果。
    首先，在頁面最上方為標題與資料來源說明，清楚標示分析檔案與使用的工作表。
    接著為「摘要」區塊，整理出資料筆數、電量 Q 的平均值、中位數、最大值與最小值，以及估算的基本電荷 e，
    讓讀者能快速掌握整體實驗結果。
    </p>

    <p>
    其次為「互動圖表」區塊，透過 Plotly 呈現 N-Q 分布圖、Q 隨次數變化圖，以及各物理量之間的關係圖，
    使電量分布與趨勢能以視覺化方式清楚呈現。再來為「主資料表」，列出每次實驗的原始與計算數據
    （如 X、T、V1、U0、Q），提供完整數據依據。
    </p>

    <p>
    最後為「Q 區間統計表」，將電量分組並統計各區間出現次數，並加上總數，使電荷的離散分布更加明顯，
    有助於判斷基本電荷的倍數關係。整體 HTML 結構條理分明，兼具數據、圖表與分析結果，
    提升報告的完整性與可讀性。
    </p>    """
    html = f"""<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>實驗五 exp5.xlsx 分析報告</title>
<style>
body {{
    font-family: Arial, 'Microsoft JhengHei', sans-serif;
    margin: 24px;
    line-height: 1.6;
    color: #222;
}}
h1, h2 {{
    margin: 0.3em 0;
}}
.card {{
    background: #fafafa;
    border: 1px solid #ddd;
    border-radius: 12px;
    padding: 16px 18px;
    margin: 16px 0;
}}
.table {{
    border-collapse: collapse;
    width: 100%;
    font-size: 14px;
}}
.table th, .table td {{
    border-bottom: 1px solid #e5e5e5;
    padding: 8px 10px;
    text-align: right;
}}
.table th {{
    background: #f3f6fb;
    text-align: center;
    position: sticky;
    top: 0;
}}
.table td:first-child, .table th:first-child {{
    text-align: center;
}}
.grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 18px;
}}
.chart {{
    min-height: 420px;
    border: 1px solid #ddd;
    border-radius: 12px;
    padding: 8px;
    background: white;
}}
.small {{
    color: #666;
    font-size: 13px;
}}
@media (max-width: 900px) {{
    .grid {{
        grid-template-columns: 1fr;
    }}
}}
</style>
</head>
<body>
<h1>密立根油滴實驗 — exp5.xlsx HTML 分析報告</h1>
<p class="small">來源檔案：{xlsx_path.name}｜使用工作表：{sheet_name}</p>

<div class="card">
  <h2>摘要</h2>
  <table class="table" style="max-width:560px">{summary_html}</table>
  <p>Q 值由公式 <code>Q = 1.945×10^-10 × V1^(3/2) / U0</code> 計算；若工作表中部分 S、V1、Q 欄位沒有快取值，程式會自動依 X、T、U0 重新補算。</p>
</div>
<div class="card">
  <h2>程式目的</h2>
  {purpose_html}
</div>
<div class="card">
  <h2>互動圖表</h2>
  {make_plotly_html(main_df, dist_df)}
</div>

<div class="card">
  <h2>主資料表</h2>
{df_to_html_table(main_df, {
    'S':'{:.6e}',
    'V1':'{:.6e}',
    'U0':'{:.1f}',
    'Q':'{:.3e}',
    'T':'{:.2f}'
})}
</div>

<div class="card">
  <h2>Q 區間統計表</h2>
  {df_to_html_table(dist_df)}
</div>
</body>
</html>
"""

    out_html.write_text(html, encoding="utf-8")
    print(f"已輸出: {out_html.resolve()}")


if __name__ == "__main__":
    main()