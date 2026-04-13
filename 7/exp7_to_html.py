# -*- coding: utf-8 -*-
# =========================================================
# 霍爾效應實驗 → python code for HTML 分析報告 from exp7.xlsx
# 功能：
# 1. 讀取 exp7.xlsx
# 2. 解析「磁場(B)」、「紅色型」、「黑色型」、「載體濃度」四個工作表
# 3. 擷取距離 d、平均磁場 B、平均電流 I、霍爾電壓 V、載體濃度 n
# 4. 進行磁場 B 與霍爾電壓 V 的線性擬合並計算 R²
# 5. 產生 Plotly 互動式圖表
# 6. 產生 HTML 表格與摘要分析
# 7. 最後輸出成 exp7_report.html
# =========================================================
from pathlib import Path
import openpyxl
import pandas as pd
import numpy as np
import json

base = Path(".")
xlsx_path = base / "exp7.xlsx"
html_path = base / "exp7_report.html"

wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# ---------------------------------------------------------
# 讀取「磁場(B)」
# ---------------------------------------------------------
ws_b = wb["磁場(B)"]
B_rows = []
for r in range(3, 8):
    d = ws_b[f"A{r}"].value
    vals = [ws_b[f"{c}{r}"].value for c in "BCDEF"]
    avg = ws_b[f"G{r}"].value
    B_rows.append([float(d), *map(float, vals), float(avg)])

df_B = pd.DataFrame(B_rows, columns=[
    "d_cm", "B1_T", "B2_T", "B3_T", "B4_T", "B5_T", "B_avg_T"
])

# ---------------------------------------------------------
# 讀取紅色型 / 黑色型
# ---------------------------------------------------------
def read_chip(ws_name):
    ws = wb[ws_name]
    rows = []
    r = 2
    while r <= 11:
        d = ws[f"A{r}"].value
        I_avg = ws[f"H{r}"].value
        V_avg = ws[f"H{r+1}"].value
        rows.append([float(d), float(I_avg), float(V_avg)])
        r += 2
    return pd.DataFrame(rows, columns=["d_cm", "I_mA", "V_mV"])

df_red = read_chip("紅色型")
df_black = read_chip("黑色型")

# ---------------------------------------------------------
# 讀取載體濃度
# ---------------------------------------------------------
ws_n = wb["載體濃度"]
rows_n = []
last_chip = None
for r in range(2, 12):
    chip = ws_n[f"A{r}"].value or last_chip
    last_chip = chip
    rows_n.append([
        chip,
        float(ws_n[f"B{r}"].value),
        float(ws_n[f"C{r}"].value),
        float(ws_n[f"D{r}"].value),
        float(ws_n[f"E{r}"].value),
        float(ws_n[f"F{r}"].value),
    ])

df_n = pd.DataFrame(rows_n, columns=["chip", "d_cm", "B_T", "I_mA", "V_mV", "n_m3"])

red_n = df_n[df_n["chip"] == "紅色型"].copy()
black_n = df_n[df_n["chip"] == "黑色型"].copy()

# ---------------------------------------------------------
# 線性擬合
# ---------------------------------------------------------
def linfit(x, y):
    coef = np.polyfit(x, y, 1)
    p = np.poly1d(coef)
    yhat = p(x)
    r2 = 1 - ((y - yhat) ** 2).sum() / ((y - y.mean()) ** 2).sum()
    return coef[0], coef[1], r2, p

red_slope, red_intercept, red_r2, red_poly = linfit(red_n["B_T"].to_numpy(), red_n["V_mV"].to_numpy())
blk_slope, blk_intercept, blk_r2, blk_poly = linfit(black_n["B_T"].to_numpy(), black_n["V_mV"].to_numpy())

# ---------------------------------------------------------
# 表格工具
# ---------------------------------------------------------
def df_to_html(df):
    return df.to_html(index=False, border=0, classes="table", justify="center")

summary_rows = [
    ("磁場量測距離範圍 (cm)", f"{df_B['d_cm'].min():.1f} ~ {df_B['d_cm'].max():.1f}"),
    ("平均磁場最大值 (T)", f"{df_B['B_avg_T'].max():.6f}"),
    ("平均磁場最小值 (T)", f"{df_B['B_avg_T'].min():.6f}"),
    ("紅色型平均電流 (mA)", f"{df_red['I_mA'].mean():.3f}"),
    ("黑色型平均電流 (mA)", f"{df_black['I_mA'].mean():.3f}"),
    ("紅色型平均載體濃度 n (m^-3)", f"{red_n['n_m3'].mean():.3e}"),
    ("黑色型平均載體濃度 n (m^-3)", f"{black_n['n_m3'].mean():.3e}"),
    ("紅色型 V-B 線性擬合斜率 (mV/T)", f"{red_slope:.3f}"),
    ("黑色型 V-B 線性擬合斜率 (mV/T)", f"{blk_slope:.3f}"),
    ("紅色型 V-B 擬合 R²", f"{red_r2:.4f}"),
    ("黑色型 V-B 擬合 R²", f"{blk_r2:.4f}"),
]

summary_html = "<table class='summary'>" + "".join(
    f"<tr><th>{k}</th><td>{v}</td></tr>" for k, v in summary_rows
) + "</table>"

purpose_html = """
<p>本程式用來將本組霍爾效應實驗資料 <code>exp7.xlsx</code> 自動整理為一份可直接在瀏覽器中閱讀的 HTML 報告。程式使用 <code>openpyxl</code> 讀取活頁簿內容，並將不同工作表中的磁場、霍爾電壓與載體濃度資料整合成結構化分析結果，方便後續整理與展示。</p>
<p>輸出的報告包含四個核心區塊：磁場(B)工作表的平均磁場整理、紅色型與黑色型半導體的霍爾電壓分析、載體濃度資料整理，以及磁場與霍爾電壓之間的線性擬合結果，並進一步建立互動式圖表與 HTML 表格。</p>
<p>互動圖表使用 Plotly 製作，可在瀏覽器中縮放、平移、滑鼠懸停查看數值，有助於更清楚比較磁場、霍爾電壓與載體濃度之間的關係，並觀察紅色型與黑色型半導體的差異。</p>
"""

# ---------------------------------------------------------
# Plotly 資料
# ---------------------------------------------------------
x_fit = np.linspace(df_n["B_T"].min(), df_n["B_T"].max(), 200)
plotly_cdn = "https://cdn.plot.ly/plotly-2.32.0.min.js"

fig1_traces = [
    {"x": df_B["d_cm"].tolist(), "y": df_B["B_avg_T"].tolist(), "type": "scatter", "mode": "lines+markers", "name": "平均磁場 B"}
]
fig1_layout = {
    "title": "圖1：距離 d 與平均磁場 B 的關係",
    "xaxis": {"title": "距離 d (cm)"},
    "yaxis": {"title": "平均磁場 B (T)"}
}

fig2_traces = [
    {"x": red_n["B_T"].tolist(), "y": red_n["V_mV"].tolist(), "type": "scatter", "mode": "markers+lines", "name": "紅色型 實驗值"},
    {"x": x_fit.tolist(), "y": red_poly(x_fit).tolist(), "type": "scatter", "mode": "lines", "name": f"紅色型 擬合線 (R²={red_r2:.4f})"},
    {"x": black_n["B_T"].tolist(), "y": black_n["V_mV"].tolist(), "type": "scatter", "mode": "markers+lines", "name": "黑色型 實驗值"},
    {"x": x_fit.tolist(), "y": blk_poly(x_fit).tolist(), "type": "scatter", "mode": "lines", "name": f"黑色型 擬合線 (R²={blk_r2:.4f})"},
]
fig2_layout = {
    "title": "圖2：磁場 B 與霍爾電壓 V 的關係",
    "xaxis": {"title": "磁場 B (T)"},
    "yaxis": {"title": "霍爾電壓 V (mV)"}
}

fig3_traces = [
    {"x": red_n["d_cm"].tolist(), "y": red_n["n_m3"].tolist(), "type": "scatter", "mode": "lines+markers", "name": "紅色型 n"},
    {"x": black_n["d_cm"].tolist(), "y": black_n["n_m3"].tolist(), "type": "scatter", "mode": "lines+markers", "name": "黑色型 n"},
]
fig3_layout = {
    "title": "圖3：距離 d 與載體濃度 n 的關係",
    "xaxis": {"title": "距離 d (cm)"},
    "yaxis": {"title": "載體濃度 n (m^-3)"}
}

fig4_traces = [
    {"x": df_red["d_cm"].tolist(), "y": df_red["V_mV"].tolist(), "type": "scatter", "mode": "lines+markers", "name": "紅色型 V"},
    {"x": df_black["d_cm"].tolist(), "y": df_black["V_mV"].tolist(), "type": "scatter", "mode": "lines+markers", "name": "黑色型 V"},
]
fig4_layout = {
    "title": "圖4：距離 d 與霍爾電壓 V 的關係",
    "xaxis": {"title": "距離 d (cm)"},
    "yaxis": {"title": "霍爾電壓 V (mV)"}
}

# ---------------------------------------------------------
# HTML 模板
# ---------------------------------------------------------
html_template = """<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>實驗七：霍爾效應 HTML 分析報告</title>
<script src="__PLOTLY_CDN__"></script>
<style>
body {
    font-family: Arial, "Microsoft JhengHei", sans-serif;
    margin: 24px;
    line-height: 1.7;
    color: #222;
    background: #fff;
}
h1, h2, h3 { margin: 0.4em 0; }
.card {
    background: #fafafa;
    border: 1px solid #ddd;
    border-radius: 12px;
    padding: 16px 18px;
    margin: 16px 0;
}
.table {
    border-collapse: collapse;
    width: 100%;
    font-size: 14px;
    background: #fff;
}
.table th, .table td {
    border-bottom: 1px solid #e5e5e5;
    padding: 8px 10px;
    text-align: right;
}
.table th {
    background: #f3f6fb;
    text-align: center;
}
.table td:first-child, .table th:first-child {
    text-align: center;
}
.summary {
    border-collapse: collapse;
    width: 100%;
    background: #fff;
}
.summary th, .summary td {
    border-bottom: 1px solid #e5e5e5;
    padding: 8px 10px;
}
.summary th { text-align: left; width: 45%; background:#f3f6fb; }
.summary td { text-align: right; }
.grid {
    display:grid;
    grid-template-columns: 1fr 1fr;
    gap: 18px;
}
.chart {
    min-height: 430px;
    border: 1px solid #ddd;
    border-radius: 12px;
    background:#fff;
}
.small { color:#666; font-size:13px; }
@media (max-width: 960px) {
    .grid { grid-template-columns: 1fr; }
}
</style>
</head>
<body>
<h1>實驗七：霍爾效應 — HTML 分析報告</h1>
<p class="small" style="font-size: 24px; font-weight: bold;">
學號：E24144773 姓名：彭靖翔
</p>
<div class="card">
  <h2>程式目的</h2>
  __PURPOSE_HTML__
</div>

<div class="card">
  <h2>摘要</h2>
  __SUMMARY_HTML__
  <p>從結果可見：距離增加時平均磁場會下降；紅色型與黑色型的霍爾電壓極性相反，顯示其主要載子型別不同；另外，霍爾電壓與磁場之間呈現近似線性關係，符合霍爾效應理論中霍爾電壓與磁場強度相關的特性。</p>
</div>

<div class="card">
  <h2>互動圖表</h2>
  <div class="grid">
    <div id="plot1" class="chart"></div>
    <div id="plot2" class="chart"></div>
  </div>
  <div class="grid" style="margin-top:18px;">
    <div id="plot3" class="chart"></div>
    <div id="plot4" class="chart"></div>
  </div>
</div>

<div class="card">
  <h2>資料表</h2>
  <h3>表1：磁場(B)工作表整理</h3>
  __TABLE_B__
  <h3>表2：紅色型量測整理</h3>
  __TABLE_RED__
  <h3>表3：黑色型量測整理</h3>
  __TABLE_BLACK__
  <h3>表4：載體濃度整理</h3>
  __TABLE_N__
</div>

<script>
Plotly.newPlot("plot1", __FIG1_TRACES__, __FIG1_LAYOUT__, {responsive:true});
Plotly.newPlot("plot2", __FIG2_TRACES__, __FIG2_LAYOUT__, {responsive:true});
Plotly.newPlot("plot3", __FIG3_TRACES__, __FIG3_LAYOUT__, {responsive:true});
Plotly.newPlot("plot4", __FIG4_TRACES__, __FIG4_LAYOUT__, {responsive:true});
</script>
</body>
</html>
"""

html = (
    html_template
    .replace("__PLOTLY_CDN__", plotly_cdn)
    .replace("__PURPOSE_HTML__", purpose_html)
    .replace("__SUMMARY_HTML__", summary_html)
    .replace("__TABLE_B__", df_to_html(df_B.round(6)))
    .replace("__TABLE_RED__", df_to_html(df_red.round(6)))
    .replace("__TABLE_BLACK__", df_to_html(df_black.round(6)))
    .replace("__TABLE_N__", df_to_html(df_n.round(6)))
    .replace("__FIG1_TRACES__", json.dumps(fig1_traces, ensure_ascii=False))
    .replace("__FIG1_LAYOUT__", json.dumps(fig1_layout, ensure_ascii=False))
    .replace("__FIG2_TRACES__", json.dumps(fig2_traces, ensure_ascii=False))
    .replace("__FIG2_LAYOUT__", json.dumps(fig2_layout, ensure_ascii=False))
    .replace("__FIG3_TRACES__", json.dumps(fig3_traces, ensure_ascii=False))
    .replace("__FIG3_LAYOUT__", json.dumps(fig3_layout, ensure_ascii=False))
    .replace("__FIG4_TRACES__", json.dumps(fig4_traces, ensure_ascii=False))
    .replace("__FIG4_LAYOUT__", json.dumps(fig4_layout, ensure_ascii=False))
)

html_path.write_text(html, encoding="utf-8")
print(f"已輸出: {html_path.resolve()}")
